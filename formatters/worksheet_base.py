"""Base worksheet helpers for Excel report generation."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class WorksheetBase:
    REPORT_TITLE = "QA COPILOT"
    REPORT_SUBTITLE = "Requirement Analysis Report"

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    TITLE_FONT = Font(bold=True, size=20)
    SUBTITLE_FONT = Font(italic=True, size=11)
    SECTION_FONT = Font(bold=True, size=14)

    # Reusable alignment for wrapped text cells
    WRAP_ALIGNMENT = Alignment(
        wrap_text=True,
        vertical="top",
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_title(self, sheet: Worksheet) -> None:
        sheet.merge_cells("A1:B1")
        sheet["A1"] = self.REPORT_TITLE
        sheet["A1"].font = self.TITLE_FONT
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("A2:B2")
        sheet["A2"] = self.REPORT_SUBTITLE
        sheet["A2"].font = self.SUBTITLE_FONT
        sheet["A2"].alignment = Alignment(horizontal="center")

    def write_table_header(self, sheet: Worksheet, row: int) -> None:
        for column, value in enumerate(("Metric", "Value"), start=1):
            cell = sheet.cell(row=row, column=column)
            cell.value = value
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    def write_metric(
        self,
        sheet: Worksheet,
        row: int,
        key: str,
        value: object,
    ) -> None:
        key_cell = sheet.cell(row=row, column=1)
        key_cell.value = key
        key_cell.font = Font(bold=True)
        key_cell.border = self.THIN_BORDER

        value_cell = sheet.cell(row=row, column=2)
        value_cell.value = value
        value_cell.border = self.THIN_BORDER
        value_cell.alignment = self.WRAP_ALIGNMENT

    def write_section(
        self,
        sheet: Worksheet,
        row: int,
        title: str,
        values: list[str],
    ) -> int:
        sheet.cell(row=row, column=1).value = title
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        if not values:
            sheet.cell(row=row, column=1).value = "-"
            sheet.cell(row=row, column=1).alignment = self.WRAP_ALIGNMENT
            return row + 2

        for value in values:
            sheet.cell(row=row, column=1).value = f"• {value}"
            sheet.cell(row=row, column=1).alignment = self.WRAP_ALIGNMENT
            row += 1

        return row + 1

    def auto_fit_columns(self, sheet: Worksheet) -> None:
        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            max_length = 0

            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=column).value
                if value:
                    max_length = max(max_length, len(str(value)))

            sheet.column_dimensions[letter].width = min(
                max(max_length + 5, 20),
                60,
            )