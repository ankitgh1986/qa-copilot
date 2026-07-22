"""Test Design worksheet."""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from .worksheet_base import WorksheetBase


class TestDesignSheet(WorksheetBase):
    """Creates the Test Design worksheet."""

    def generate(self, sheet: Worksheet, result) -> None:
        """Populate the Test Design worksheet."""

        self.write_title(sheet)

        design = result.test_design

        row = 4

        # =====================================================
        # Requirement Understanding
        # =====================================================

        sheet.cell(row=row, column=1).value = "Requirement Understanding"
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        self.write_table_header(sheet, row)
        row += 1

        self.write_metric(sheet, row, "Domain", design.domain)
        row += 1

        self.write_metric(sheet, row, "Feature", design.feature_name)
        row += 1

        self.write_metric(sheet, row, "Business Goal", design.business_goal)
        row += 2

        row = self.write_section(
            sheet,
            row,
            "Actors",
            design.actors,
        )

        row = self.write_section(
            sheet,
            row,
            "Business Workflow",
            design.business_workflow,
        )

        row = self.write_section(
            sheet,
            row,
            "Business Rules",
            design.business_rules,
        )

        row = self.write_section(
            sheet,
            row,
            "Dependencies",
            design.dependencies,
        )

        row = self.write_section(
            sheet,
            row,
            "Integration Points",
            design.integration_points,
        )

        row = self.write_section(
            sheet,
            row,
            "Assumptions",
            design.assumptions,
        )

        # =====================================================
        # QA Analysis
        # =====================================================

        row = self.write_section(
            sheet,
            row,
            "Test Objectives",
            design.test_objectives,
        )

        row = self.write_section(
            sheet,
            row,
            "Quality Attributes",
            design.quality_attributes,
        )

        row = self.write_section(
            sheet,
            row,
            "Functional Areas",
            design.functional_areas,
        )

        row = self.write_section(
            sheet,
            row,
            "Applicable Test Types",
            design.applicable_test_types,
        )

        row = self.write_section(
            sheet,
            row,
            "Coverage Strategy",
            design.coverage_strategy,
        )

        row = self.write_section(
            sheet,
            row,
            "Test Data Requirements",
            design.test_data_requirements,
        )

        # =====================================================
        # Identified Risks
        # =====================================================

        sheet.cell(row=row, column=1).value = "Identified Risks"
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        headers = [
            "Risk",
            "Category",
            "Impact",
        ]

        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER

        row += 1

        for risk in design.identified_risks:
            sheet.cell(row=row, column=1).value = risk.risk
            sheet.cell(row=row, column=2).value = risk.category
            sheet.cell(row=row, column=3).value = risk.impact

            for column in range(1, 4):
                sheet.cell(row=row, column=column).border = self.THIN_BORDER

            row += 1

        row += 1

        # =====================================================
        # Automation Strategy
        # =====================================================

        row = self.write_section(
            sheet,
            row,
            "API Automation",
            design.automation_candidates.api_automation,
        )

        row = self.write_section(
            sheet,
            row,
            "UI Automation",
            design.automation_candidates.ui_automation,
        )

        row = self.write_section(
            sheet,
            row,
            "Performance Automation",
            design.automation_candidates.performance_automation,
        )

        row = self.write_section(
            sheet,
            row,
            "Manual Exploratory",
            design.automation_candidates.manual_exploratory,
        )

        row = self.write_section(
            sheet,
            row,
            "Automation Strategy",
            design.automation_strategy,
        )

        sheet.freeze_panes = "A5"

        self.auto_fit_columns(sheet)