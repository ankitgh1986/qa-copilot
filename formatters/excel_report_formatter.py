"""Excel report formatter for QA Copilot."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.requirement_analyzer import RequirementAnalysisResult
from .acceptance_criteria_sheet import AcceptanceCriteriaSheet
from .ambiguity_assessment_sheet import AmbiguityAssessmentSheet
from .executive_summary_sheet import ExecutiveSummarySheet
from .quality_assessment_sheet import QualityAssessmentSheet
from .requirement_summary_sheet import RequirementSummarySheet
from .risk_assessment_sheet import RiskAssessmentSheet
from .test_case_sheet import TestCasesSheet
from .test_design_sheet import TestDesignSheet
from .improvement_assessment import ImprovementAssessmentSheet
from .worksheet_base import WorksheetBase

logger = logging.getLogger(__name__)


class ExcelReportFormatter:
    """Produce an Excel workbook from requirement analysis results."""

    _WORKSHEETS: list[tuple[str, type[WorksheetBase]]] = [
        ("Executive Summary", ExecutiveSummarySheet),
        ("Requirement Summary", RequirementSummarySheet),
        ("Quality Assessment", QualityAssessmentSheet),
        ("Ambiguity Assessment", AmbiguityAssessmentSheet),
        ("Risk Assessment", RiskAssessmentSheet),
        ("Improvement Assessment", ImprovementAssessmentSheet),
        ("Acceptance Criteria", AcceptanceCriteriaSheet),
        ("Test Design", TestDesignSheet),
        ("Test Cases", TestCasesSheet),
    ]

    def format(
        self,
        result: RequirementAnalysisResult,
        output_path: str | Path,
    ) -> None:
        """Generate and save the Excel report workbook.

        Args:
            result: The requirement analysis result to render.
            output_path: The target path for the generated workbook.
        """

        logger.info(
            "Starting Excel report generation for output: %s",
            output_path,
        )

        workbook = self._create_workbook()

        try:
            self._populate_workbook(workbook, result)
            self._save_workbook(workbook, output_path)

            logger.info(
                "Excel report generated successfully at %s",
                output_path,
            )

        except Exception:
            logger.exception("Failed to generate Excel report.")
            raise

    def _create_workbook(self) -> Workbook:
        """Create a new openpyxl workbook."""

        return Workbook()

    def _populate_workbook(
        self,
        workbook: Workbook,
        result: RequirementAnalysisResult,
    ) -> None:
        """Add report worksheets to the workbook."""

        for title, worksheet_class in self._WORKSHEETS:
            worksheet = workbook.create_sheet(title=title)
            worksheet_class().generate(worksheet, result)

        self._cleanup_default_sheet(workbook)

    def _cleanup_default_sheet(self, workbook: Workbook) -> None:
        """Remove the default workbook sheet if it was not used."""

        default_title = "Sheet"

        if default_title in workbook.sheetnames:
            workbook.remove(workbook[default_title])

    def _save_workbook(
        self,
        workbook: Workbook,
        output_path: str | Path,
    ) -> None:
        """Save the workbook to disk, creating parent directories as needed."""

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_file)
