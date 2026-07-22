"""Ambiguity Assessment worksheet."""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from .worksheet_base import WorksheetBase


class AmbiguityAssessmentSheet(WorksheetBase):
    """Creates the Ambiguity Assessment worksheet."""

    def generate(self, sheet: Worksheet, result) -> None:
        """Populate the Ambiguity Assessment worksheet."""

        self.write_title(sheet)

        ambiguity = result.ambiguity_assessment

        row = 4

        # ---------------------------------------------------------
        # Metrics
        # ---------------------------------------------------------

        sheet.cell(row=row, column=1).value = "Ambiguity Metrics"
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        self.write_table_header(sheet, row)
        row += 1

        metrics = [
            ("Ambiguity Score", ambiguity.ambiguity_score),
            ("Ambiguity Level", ambiguity.ambiguity_level),
        ]

        for key, value in metrics:
            self.write_metric(sheet, row, key, value)
            row += 1

        row += 2

        # ---------------------------------------------------------
        # Findings
        # ---------------------------------------------------------

        row = self.write_section(
            sheet,
            row,
            "Ambiguous Terms",
            ambiguity.ambiguous_terms,
        )

        row = self.write_section(
            sheet,
            row,
            "Missing Information",
            ambiguity.missing_information,
        )

        row = self.write_section(
            sheet,
            row,
            "Undefined Business Rules",
            ambiguity.undefined_business_rules,
        )

        row = self.write_section(
            sheet,
            row,
            "Assumptions",
            ambiguity.assumptions,
        )

        row = self.write_section(
            sheet,
            row,
            "Conflicting Statements",
            ambiguity.conflicting_statements,
        )

        # ---------------------------------------------------------
        # Summary
        # ---------------------------------------------------------

        sheet.cell(row=row, column=1).value = "Summary"
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row + 4,
            end_column=2,
        )

        summary = sheet.cell(row=row, column=1)
        summary.value = ambiguity.summary
        summary.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        sheet.row_dimensions[row].height = 100

        sheet.freeze_panes = "A5"

        self.auto_fit_columns(sheet)