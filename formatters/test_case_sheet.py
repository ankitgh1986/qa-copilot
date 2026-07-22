"""Test Cases worksheet."""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from .worksheet_base import WorksheetBase


class TestCasesSheet(WorksheetBase):
    """Creates the Test Cases worksheet."""

    def generate(self, sheet: Worksheet, result) -> None:
        """Populate the Test Cases worksheet."""

        self.write_title(sheet)

        # Result contains a TestSuite object, not a list of test cases.
        test_suite = result.test_cases

        row = 4

        for index, test_case in enumerate(
            test_suite.test_cases,
            start=1,
        ):

            # =====================================================
            # Test Case Header
            # =====================================================

            sheet.cell(row=row, column=1).value = (
                f"Test Case {index} - {test_case.test_case_id}"
            )
            sheet.cell(row=row, column=1).font = self.SECTION_FONT
            row += 1

            self.write_table_header(sheet, row)
            row += 1

            metrics = [
                ("Title", test_case.title),
                ("Objective", test_case.objective),
                ("Priority", test_case.priority),
                ("Severity", test_case.severity),
                ("Risk", test_case.risk),
                ("Automation Candidate", test_case.automation_candidate),
                ("Automation Layer", test_case.automation_layer),
            ]

            for key, value in metrics:
                self.write_metric(sheet, row, key, value)
                row += 1

            row += 1

            # =====================================================
            # Preconditions
            # =====================================================

            row = self.write_section(
                sheet,
                row,
                "Preconditions",
                test_case.preconditions,
            )

            # =====================================================
            # Test Data
            # =====================================================

            row = self.write_section(
                sheet,
                row,
                "Test Data",
                test_case.test_data,
            )

            # =====================================================
            # Tags
            # =====================================================

            row = self.write_section(
                sheet,
                row,
                "Tags",
                test_case.tags,
            )

            # =====================================================
            # Test Steps
            # =====================================================

            sheet.cell(row=row, column=1).value = "Test Steps"
            sheet.cell(row=row, column=1).font = self.SECTION_FONT
            row += 1

            headers = [
                "Step",
                "Action",
                "Expected Result",
            ]

            for column, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row, column=column)
                cell.value = header
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.THIN_BORDER
                cell.alignment = self.WRAP_ALIGNMENT

            row += 1

            for step in test_case.test_steps:

                sheet.cell(row=row, column=1).value = step.step_number
                sheet.cell(row=row, column=2).value = step.action
                sheet.cell(row=row, column=3).value = step.expected_result

                for column in range(1, 4):
                    cell = sheet.cell(row=row, column=column)
                    cell.border = self.THIN_BORDER
                    cell.alignment = self.WRAP_ALIGNMENT

                row += 1

            row += 2

        sheet.freeze_panes = "A5"

        self.auto_fit_columns(sheet)