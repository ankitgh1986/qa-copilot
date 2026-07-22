"""Requirement Summary worksheet."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from .worksheet_base import WorksheetBase


class RequirementSummarySheet(WorksheetBase):
    """Creates the Requirement Summary worksheet."""

    def generate(self, sheet: Worksheet, result) ->None:
        """Populate the Requirement Summary worksheet."""

        self.write_title(sheet)

        row = 4

        # ------------------------------------------------------------------
        # Document Information
        # ------------------------------------------------------------------

        sheet.cell(row=row, column=1).value = "Document Information"
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        self.write_table_header(sheet, row)
        row += 1

        context = result.context

        metrics = [
            ("Source File", getattr(context, "source_file", "-")),
            ("Document Type", getattr(context, "document_type", "-")),
            ("Language", getattr(context, "language", "-")),
            ("Pages", getattr(context, "page_count", "-")),
            ("Words", getattr(context, "word_count", "-")),
        ]

        for key, value in metrics:
            self.write_metric(sheet, row, key, value)
            row += 1

        row += 2

        # ------------------------------------------------------------------
        # Requirement Summary
        # ------------------------------------------------------------------

        sheet.cell(row=row, column=1).value = "Requirement Summary"
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        summary_cell = sheet.cell(row=row, column=1)
        summary_cell.value = context.short_summary
        summary_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        sheet.row_dimensions[row].height = 160

        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row + 5,
            end_column=2,
        )

        row += 7

        # ------------------------------------------------------------------
        # Functional Areas
        # ------------------------------------------------------------------

        functional_areas = getattr(
            context,
            "functional_areas",
            [],
        )

        row = self.write_section(
            sheet,
            row,
            "Functional Areas",
            functional_areas,
        )

        # ------------------------------------------------------------------
        # Business Objectives
        # ------------------------------------------------------------------

        objectives = getattr(
            context,
            "business_objectives",
            [],
        )

        row = self.write_section(
            sheet,
            row,
            "Business Objectives",
            objectives,
        )

        # ------------------------------------------------------------------
        # Key Features
        # ------------------------------------------------------------------

        features = getattr(
            context,
            "key_features",
            [],
        )

        row = self.write_section(
            sheet,
            row,
            "Key Features",
            features,
        )

        # ------------------------------------------------------------------
        # Stakeholders
        # ------------------------------------------------------------------

        stakeholders = getattr(
            context,
            "stakeholders",
            [],
        )

        row = self.write_section(
            sheet,
            row,
            "Stakeholders",
            stakeholders,
        )

        # ------------------------------------------------------------------
        # Assumptions
        # ------------------------------------------------------------------

        assumptions = getattr(
            context,
            "assumptions",
            [],
        )

        self.write_section(
            sheet,
            row,
            "Assumptions",
            assumptions,
        )

        sheet.freeze_panes = "A5"

        self.auto_fit_columns(sheet)