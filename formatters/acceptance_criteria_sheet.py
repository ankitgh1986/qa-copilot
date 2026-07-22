"""Acceptance Criteria worksheet."""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from .worksheet_base import WorksheetBase


class AcceptanceCriteriaSheet(WorksheetBase):
    """Creates the Acceptance Criteria worksheet."""

    def generate(self, sheet: Worksheet, result) -> None:
        """Populate the Acceptance Criteria worksheet."""

        self.write_title(sheet)

        acceptance = result.acceptance_criteria

        row = 4

        # ---------------------------------------------------------
        # Section Header
        # ---------------------------------------------------------

        sheet.cell(row=row, column=1).value = "Acceptance Criteria"
        sheet.cell(row=row, column=1).font = self.SECTION_FONT
        row += 1

        # ---------------------------------------------------------
        # Table Header
        # ---------------------------------------------------------

        headers = [
            "No.",
            "Acceptance Criterion",
        ]

        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER

        row += 1

        criteria = acceptance.criteria or []

        if not criteria:
            sheet.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=2,
            )

            cell = sheet.cell(row=row, column=1)
            cell.value = "No acceptance criteria identified."
            cell.border = self.THIN_BORDER

        else:
            for index, criterion in enumerate(criteria, start=1):

                number_cell = sheet.cell(row=row, column=1)
                number_cell.value = index
                number_cell.border = self.THIN_BORDER

                criterion_cell = sheet.cell(row=row, column=2)
                criterion_cell.value = criterion
                criterion_cell.border = self.THIN_BORDER
                criterion_cell.alignment = self.WRAP_ALIGNMENT

                row += 1

        sheet.freeze_panes = "A6"

        self.auto_fit_columns(sheet)